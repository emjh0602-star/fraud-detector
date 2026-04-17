from flask import Flask, render_template, request, jsonify, send_file, session, redirect
import pandas as pd
import numpy as np
import json, os, io, re, hashlib, secrets
from datetime import datetime, timedelta
from collections import defaultdict
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.permanent_session_lifetime = timedelta(hours=8)

DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)
HISTORY_FILE = os.path.join(DATA_DIR, "history.json")
USERS_FILE   = os.path.join(DATA_DIR, "users.json")
CORPS_FILE   = os.path.join(DATA_DIR, "corps.json")
AUDIT_FILE   = os.path.join(DATA_DIR, "audit.log")

# ── 유틸 ──────────────────────────────────────────────
def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, encoding="utf-8") as f:
            return json.load(f)
    default = {"admin": {"password": hash_pw("admin1234!"), "name": "관리자",
                         "role": "admin", "created_at": datetime.now().isoformat()}}
    save_users(default)
    return default

def save_users(u):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(u, f, ensure_ascii=False, indent=2)

def load_corps():
    if os.path.exists(CORPS_FILE):
        with open(CORPS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return []  # [{"id": "corp_001", "name": "(주)ABC코리아", "manager": "hong"}]

def save_corps(corps):
    with open(CORPS_FILE, "w", encoding="utf-8") as f:
        json.dump(corps, f, ensure_ascii=False, indent=2)

def audit(action, detail=""):
    line = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {session.get('user','?')} | {action} | {detail}\n"
    with open(AUDIT_FILE, "a", encoding="utf-8") as f:
        f.write(line)

def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_history(h):
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(h, f, ensure_ascii=False, indent=2)

# ── 인증 데코레이터 ───────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "로그인이 필요합니다.", "redirect": "/login"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") != "admin":
            return jsonify({"error": "관리자 권한이 필요합니다."}), 403
        return f(*args, **kwargs)
    return decorated

# ── 페이지 라우트 ─────────────────────────────────────
@app.route("/")
def index():
    if "user" not in session:
        return redirect("/login")
    return render_template("index.html", username=session["name"],
                           role=session["role"], userid=session["user"])

@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "GET":
        return render_template("login.html") if "user" not in session else redirect("/")
    d = request.get_json() or {}
    uid, pw = d.get("username","").strip(), d.get("password","")
    users = load_users()
    user  = users.get(uid)
    if not user or user["password"] != hash_pw(pw):
        audit("LOGIN_FAIL", uid)
        return jsonify({"error": "아이디 또는 비밀번호가 올바르지 않습니다."}), 401
    session.permanent = True
    session["user"] = uid
    session["name"] = user["name"]
    session["role"] = user["role"]
    audit("LOGIN_OK")
    return jsonify({"ok": True, "role": user["role"], "name": user["name"]})

@app.route("/logout")
def logout():
    audit("LOGOUT")
    session.clear()
    return redirect("/login")

# ── 법인 관리 API ─────────────────────────────────────
@app.route("/api/corps", methods=["GET"])
@login_required
def get_corps():
    corps = load_corps()
    uid = session["user"]
    role = session["role"]
    # 관리자는 전체, 일반 사용자는 담당 법인만
    if role == "admin":
        return jsonify(corps)
    else:
        return jsonify([c for c in corps if c.get("manager") == uid])

@app.route("/api/corps", methods=["POST"])
@login_required
@admin_required
def create_corp():
    d = request.get_json() or {}
    name = d.get("name","").strip()
    manager = d.get("manager","").strip()
    if not name:
        return jsonify({"error": "법인명을 입력해주세요."}), 400
    corps = load_corps()
    if any(c["name"] == name for c in corps):
        return jsonify({"error": "이미 등록된 법인명입니다."}), 409
    corp_id = f"corp_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    corps.append({"id": corp_id, "name": name, "manager": manager,
                  "created_at": datetime.now().isoformat()})
    save_corps(corps)
    audit("CORP_CREATE", f"{name} → {manager}")
    return jsonify({"ok": True, "message": f"'{name}' 법인이 등록되었습니다."})

@app.route("/api/corps/<corp_id>", methods=["PUT"])
@login_required
@admin_required
def update_corp(corp_id):
    d = request.get_json() or {}
    corps = load_corps()
    for c in corps:
        if c["id"] == corp_id:
            c["name"] = d.get("name", c["name"]).strip()
            c["manager"] = d.get("manager", c["manager"]).strip()
            save_corps(corps)
            audit("CORP_UPDATE", c["name"])
            return jsonify({"ok": True})
    return jsonify({"error": "법인을 찾을 수 없습니다."}), 404

@app.route("/api/corps/<corp_id>", methods=["DELETE"])
@login_required
@admin_required
def delete_corp(corp_id):
    corps = load_corps()
    target = next((c for c in corps if c["id"] == corp_id), None)
    if not target:
        return jsonify({"error": "법인을 찾을 수 없습니다."}), 404
    corps = [c for c in corps if c["id"] != corp_id]
    save_corps(corps)
    # 관련 학습 데이터도 삭제
    h = load_history()
    for key in [target["name"], target["name"]+"__cumul__"]:
        if key in h: del h[key]
    save_history(h)
    audit("CORP_DELETE", target["name"])
    return jsonify({"ok": True, "message": f"'{target['name']}' 법인이 삭제되었습니다."})

# ── 분석 API ──────────────────────────────────────────
@app.route("/api/upload", methods=["POST"])
@login_required
def upload():
    files     = request.files.getlist("file")  # 여러 파일 지원
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "파일이 없습니다."}), 400
    corp_id   = request.form.get("corp_id","").strip()
    file_type = request.form.get("file_type","new")

    # 법인 검증
    corps = load_corps()
    corp  = next((c for c in corps if c["id"] == corp_id), None)
    if not corp:
        return jsonify({"error": "법인을 선택해주세요."}), 400

    # 담당자 확인 (관리자는 전체 가능)
    if session["role"] != "admin" and corp.get("manager") != session["user"]:
        return jsonify({"error": "담당 법인이 아닙니다."}), 403

    corp_name = corp["name"]

    # 여러 파일 합치기
    all_rows = []
    for file in files:
        if file.filename == "": continue
        try:
            fname = file.filename.lower()
            # 스마트 파서 먼저 시도 (xlsx/xls)
            if fname.endswith(('.xlsx', '.xls')):
                parsed = smart_parse_excel(file, file.filename)
                if parsed:
                    for r in parsed:
                        r['_source_file'] = file.filename
                    all_rows.extend(parsed)
                else:
                    # 스마트 파서 실패시 pandas로 폴백
                    file.seek(0)
                    df = pd.read_excel(file)
                    df = normalize_columns(df).fillna("")
                    df['_source_file'] = file.filename
                    all_rows.extend([{k: str(v) if not isinstance(v,(int,float)) else v
                                      for k,v in r.items()} for r in df.to_dict("records")])
            else:
                # CSV
                df = pd.read_csv(file, encoding="utf-8-sig")
                df = normalize_columns(df).fillna("")
                df['_source_file'] = file.filename
                all_rows.extend([{k: str(v) if not isinstance(v,(int,float)) else v
                                  for k,v in r.items()} for r in df.to_dict("records")])
        except Exception as e:
            return jsonify({"error": f"파일 읽기 오류 ({file.filename}): {e}"}), 400

    if not all_rows:
        return jsonify({"error": "읽을 수 있는 데이터가 없습니다."}), 400

    rows = all_rows
    file_count = len(files)

    if file_type == "history":
        h = load_history()
        existing = h.get(corp_name, [])
        existing_payees = {r.get("payee","") for r in existing if r.get("payee")}
        if not existing:
            h[corp_name] = rows
            saved_count = len(rows)
        else:
            new_rows = [r for r in rows if r.get("payee","") not in existing_payees]
            h[corp_name] = existing + new_rows
            saved_count = len(new_rows)
        save_history(h)
        total_count = len(h[corp_name])
        audit("HISTORY_SAVE", f"{corp_name} +{saved_count}건 (누적 {total_count}건)")
        return jsonify({
            "message": f"'{corp_name}' 패턴 학습 완료 · 신규 {saved_count}건 추가 (누적 {total_count}건)",
            "count": saved_count,
            "total": total_count,
            "type": "history"
        })

    h       = load_history()
    results = analyze_transactions(rows, corp_name, h)
    danger  = sum(1 for r in results if r["risk"]=="이상")
    audit("ANALYZE", f"{corp_name} {len(results)}건 ({file_count}개 파일) → 이상{danger}")

    # 누적 학습
    cumul_key = corp_name + "__cumul__"
    existing_cumul = h.get(cumul_key, [])
    existing_payees = {r.get("payee","") for r in existing_cumul}
    new_rows = [{"payee": r.get("payee",""), "amount": r.get("amount",""),
                 "account": r.get("account",""), "date": r.get("date","")}
                for r in rows if r.get("payee","") not in existing_payees and r.get("payee","")]
    h[cumul_key] = existing_cumul + new_rows
    save_history(h)

    return jsonify({
        "type": "analysis", "corp_name": corp_name,
        "total": len(results),
        "danger": danger,
        "warning": sum(1 for r in results if r["risk"]=="주의"),
        "ok":     sum(1 for r in results if r["risk"]=="정상"),
        "total_amount": sum(r["amount_clean"] for r in results),
        "has_history": bool(h.get(corp_name)),
        "cumul_payees": len(h[cumul_key]),
        "results": results,
        "analyzed_by": session["name"]
    })

@app.route("/api/export", methods=["POST"])
@login_required
def export_csv():
    data = request.json.get("data",[])
    if not data: return jsonify({"error": "데이터 없음"}), 400
    rows = [{"법인명":r.get("corp",""),"거래일자":r.get("date",""),"수취인명":r.get("payee",""),
             "금액":r.get("amount",""),"계좌번호":r.get("account",""),"적요":r.get("memo",""),
             "위험도":r.get("risk",""),"감지사유":" / ".join(r.get("reasons",[])),"위험점수":r.get("risk_score",0)}
            for r in data]
    buf = io.BytesIO()
    pd.DataFrame(rows).to_csv(buf, index=False, encoding="utf-8-sig")
    buf.seek(0)
    audit("EXPORT_CSV", f"{len(data)}건")
    return send_file(buf, mimetype="text/csv", as_attachment=True,
                     download_name=f"이상거래분석_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")

@app.route("/api/history/<corp_name>", methods=["DELETE"])
@login_required
def delete_history(corp_name):
    h = load_history()
    for key in [corp_name, corp_name+"__cumul__"]:
        if key in h: del h[key]
    save_history(h)
    audit("HISTORY_DELETE", corp_name)
    return jsonify({"message": f"'{corp_name}' 학습 데이터 삭제 완료"})

# ── 관리자 전용 API ────────────────────────────────────
@app.route("/api/admin/users", methods=["GET"])
@login_required
@admin_required
def list_users():
    users = load_users()
    return jsonify([{"id":uid,"name":u["name"],"role":u["role"],"created_at":u.get("created_at","")}
                    for uid,u in users.items()])

@app.route("/api/admin/users", methods=["POST"])
@login_required
@admin_required
def create_user():
    d = request.get_json() or {}
    uid,pw,name,role = d.get("username","").strip(),d.get("password","").strip(),d.get("name","").strip(),d.get("role","user")
    if not uid or not pw or not name:
        return jsonify({"error": "아이디, 이름, 비밀번호를 모두 입력해주세요."}), 400
    if len(pw) < 8:
        return jsonify({"error": "비밀번호는 8자 이상이어야 합니다."}), 400
    users = load_users()
    if uid in users: return jsonify({"error": "이미 존재하는 아이디입니다."}), 409
    users[uid] = {"password": hash_pw(pw), "name": name, "role": role,
                  "created_at": datetime.now().isoformat()}
    save_users(users)
    audit("USER_CREATE", f"{uid} ({name})")
    return jsonify({"ok": True, "message": f"'{name}' 계정이 생성되었습니다."})

@app.route("/api/admin/users/<uid>", methods=["DELETE"])
@login_required
@admin_required
def delete_user(uid):
    if uid == session["user"]: return jsonify({"error": "자신의 계정은 삭제할 수 없습니다."}), 400
    users = load_users()
    if uid not in users: return jsonify({"error": "존재하지 않는 계정입니다."}), 404
    name = users[uid]["name"]; del users[uid]; save_users(users)
    audit("USER_DELETE", f"{uid} ({name})")
    return jsonify({"ok": True, "message": f"'{name}' 계정이 삭제되었습니다."})

@app.route("/api/admin/users/<uid>/password", methods=["PUT"])
@login_required
@admin_required
def reset_password(uid):
    d = request.get_json() or {}
    pw = d.get("password","").strip()
    if len(pw) < 8: return jsonify({"error": "비밀번호는 8자 이상이어야 합니다."}), 400
    users = load_users()
    if uid not in users: return jsonify({"error": "존재하지 않는 계정입니다."}), 404
    users[uid]["password"] = hash_pw(pw); save_users(users)
    audit("PW_RESET", uid)
    return jsonify({"ok": True, "message": "비밀번호가 변경되었습니다."})

@app.route("/api/admin/audit")
@login_required
@admin_required
def get_audit():
    if not os.path.exists(AUDIT_FILE): return jsonify({"logs": []})
    with open(AUDIT_FILE, encoding="utf-8") as f:
        lines = f.readlines()[-200:]
    return jsonify({"logs": [l.rstrip() for l in reversed(lines)]})

@app.route("/api/me/password", methods=["PUT"])
@login_required
def change_own_password():
    d = request.get_json() or {}
    old_pw,new_pw = d.get("old_password",""),d.get("new_password","").strip()
    users = load_users(); me = users.get(session["user"])
    if not me or me["password"] != hash_pw(old_pw):
        return jsonify({"error": "현재 비밀번호가 올바르지 않습니다."}), 401
    if len(new_pw) < 8: return jsonify({"error": "새 비밀번호는 8자 이상이어야 합니다."}), 400
    users[session["user"]]["password"] = hash_pw(new_pw); save_users(users)
    audit("PW_CHANGE_SELF")
    return jsonify({"ok": True, "message": "비밀번호가 변경되었습니다."})

# ── 분석 로직 ──────────────────────────────────────────

SKIP_KEYWORDS = ['합계','소계','총계','총 계','총현금','*','**','경비','정산','이체','예금','현금','소계','계좌명','합 계']

def smart_parse_excel(file_obj, filename):
    """복잡한 엑셀 구조도 자동으로 파싱하는 스마트 파서"""
    import openpyxl
    from io import BytesIO

    file_obj.seek(0)
    wb = openpyxl.load_workbook(BytesIO(file_obj.read()), data_only=True)
    ws = wb.active

    # 헤더 행 자동 탐지 (키워드 매칭)
    header_keywords = ['거래처','수취인','금액','계좌','적요','계정','은행','날짜','일자','이체','거래처명','원화금액','계정과목','지급금액']
    header_row = None
    best_row = None
    best_matches = 0
    for i in range(1, min(60, ws.max_row+1)):
        row_vals = [str(ws.cell(i,c).value or '') for c in range(1, min(20, ws.max_column+1))]
        matches = sum(1 for k in header_keywords if any(k in v for v in row_vals))
        if matches > best_matches:
            best_matches = matches
            best_row = i
        if matches >= 5:
            header_row = i
            break
    if not header_row and best_matches >= 3:
        header_row = best_row

    if not header_row:
        # 헤더 못 찾으면 첫 번째 행을 헤더로 사용
        header_row = 1

    # 헤더 컬럼 매핑
    headers = {}
    for c in range(1, ws.max_column+1):
        val = str(ws.cell(header_row, c).value or '').strip()
        if not val: continue
        v = val.lower().replace(' ','')
        if any(k in v for k in ['일자','date','날짜','거래일']): headers[c] = 'date'
        elif any(k in v for k in ['수취인','거래처','payee','받는','상대방','입금처','업체','예금주','거래처명']): 
            if 'payee' not in headers.values(): headers[c] = 'payee'
        elif any(k in v for k in ['금액','amount','지급액','출금','이체금액','원화금액','지급금액']):
            if 'amount' not in headers.values(): headers[c] = 'amount'
        elif any(k in v for k in ['계좌번호','account','계좌']):
            if 'account' not in headers.values(): headers[c] = 'account'
        elif any(k in v for k in ['은행','bank']):
            if 'bank' not in headers.values(): headers[c] = 'bank'
        elif any(k in v for k in ['적요','memo','내용','비고','remark','거래내용','계정','비용','계정과목']):
            if 'memo' not in headers.values(): headers[c] = 'memo'

    # 역방향 매핑 (첫번째 매칭만 사용)
    col_map = {}
    used = set()
    for c, field in headers.items():
        if field not in used:
            col_map[c] = field
            used.add(field)

    # 수취인 컬럼이 없으면 두 번째로 긴 텍스트 컬럼 추정
    if 'payee' not in col_map.values():
        for c in range(1, ws.max_column+1):
            val = str(ws.cell(header_row, c).value or '').strip()
            if val and c not in col_map:
                col_map[c] = 'payee'
                break

    # 데이터 추출 (합계/섹션 제목 행 건너뜀)
    rows = []
    for i in range(header_row+1, ws.max_row+1):
        # 수취인 컬럼 확인
        payee_col = next((c for c,f in col_map.items() if f=='payee'), None)
        amt_col   = next((c for c,f in col_map.items() if f=='amount'), None)

        payee_val = str(ws.cell(i, payee_col).value or '').strip() if payee_col else ''
        amt_val   = ws.cell(i, amt_col).value if amt_col else None

        if not payee_val: continue
        # 합계/섹션 행 건너뜀
        if any(k in payee_val for k in SKIP_KEYWORDS): continue
        if payee_val in ['거래처','수취인','업체명']: continue

        # 금액 검증 (숫자여야 함)
        try:
            amt_num = float(re.sub(r'[^0-9.]', '', str(amt_val))) if amt_val else 0
        except:
            amt_num = 0

        row = {'payee': payee_val}
        for c, field in col_map.items():
            val = ws.cell(i, c).value
            if field == 'amount':
                row['amount'] = amt_num
            elif field == 'account':
                # 계좌번호: 은행명과 합치기
                bank_col = next((bc for bc,bf in col_map.items() if bf=='bank'), None)
                bank = str(ws.cell(i, bank_col).value or '').strip() if bank_col else ''
                acct = str(val or '').strip()
                row['account'] = f"{bank} {acct}".strip() if bank else acct
            elif field not in ['payee','bank']:
                row[field] = str(val or '').strip()

        rows.append(row)

    return rows

def normalize_columns(df):
    """기존 pandas DataFrame 방식 - 단순 구조 엑셀용"""
    col_map = {}
    for col in df.columns:
        c = str(col).strip().lower().replace(" ","")
        if any(k in c for k in ["일자","date","날짜","거래일"]): col_map[col]="date"
        elif any(k in c for k in ["수취인","payee","받는","거래처","상대방","입금처","업체","예금주"]): 
            if "payee" not in col_map.values(): col_map[col]="payee"
        elif any(k in c for k in ["금액","amount","amt","지급액","출금","이체금액"]):
            if "amount" not in col_map.values(): col_map[col]="amount"
        elif any(k in c for k in ["계좌번호","계좌","account","bank","은행"]):
            if "account" not in col_map.values(): col_map[col]="account"
        elif any(k in c for k in ["적요","memo","내용","비고","remark","거래내용","계정","비용구분"]):
            if "memo" not in col_map.values(): col_map[col]="memo"
    return df.rename(columns=col_map)

def clean_amount(val):
    s = re.sub(r"[^0-9.]","",str(val))
    try: return float(s)
    except: return 0.0

def analyze_transactions(rows, corp_name, history):
    corp_history = history.get(corp_name, [])
    cumul_key    = corp_name + "__cumul__"
    cumul_data   = history.get(cumul_key, [])
    all_history  = corp_history + cumul_data

    hist_amounts = [a for a in (clean_amount(r.get("amount",0)) for r in all_history) if a>0]
    hist_payees  = {r.get("payee","") for r in all_history if r.get("payee")}
    has_any_history = len(all_history) > 0

    avg_amt = np.mean(hist_amounts) if hist_amounts else 0
    p95_amt = np.percentile(hist_amounts,95) if hist_amounts else 0

    payee_count, amount_count = defaultdict(int), defaultdict(int)
    for r in rows:
        if r.get("payee"):  payee_count[r["payee"]] += 1
        if r.get("amount"): amount_count[str(r["amount"])] += 1

    results = []
    for r in rows:
        reasons, score = [], 0
        amt = clean_amount(r.get("amount",0))

        if amt>=500_000_000: reasons.append("5억원 이상 초고액"); score+=3
        elif amt>=100_000_000: reasons.append("1억원 이상 고액"); score+=2
        elif amt>=50_000_000: reasons.append("5천만원 이상"); score+=1

        if avg_amt>0 and amt>avg_amt*5: reasons.append(f"평균 대비 {int(amt/avg_amt)}배 초과"); score+=2
        if p95_amt>0 and amt>p95_amt*2: reasons.append("상위 5% 금액의 2배 초과"); score+=2

        payee = r.get("payee","")
        if payee and has_any_history and payee not in hist_payees:
            reasons.append("신규 거래처"); score+=1
            if amt>=10_000_000: reasons.append("신규 거래처 고액"); score+=1
        if payee and not has_any_history:
            reasons.append("거래처 이력 없음 (패턴 학습 필요)")

        if payee_count.get(payee,0)>=3: reasons.append(f"동일 수취인 {payee_count[payee]}건 반복"); score+=1
        if re.search(r"긴급|급건|urgent|즉시|당일처리",str(r.get("memo","")),re.I): reasons.append("긴급 처리 요청"); score+=1
        if amount_count.get(str(r.get("amount","")),0)>=3 and amt>0: reasons.append("동일 금액 3건 이상 반복"); score+=1
        acct = str(r.get("account","")).strip()
        if (not acct or acct in ["nan","-",""]) and amt>=1_000_000: reasons.append("계좌번호 미기재"); score+=1

        risk = "이상" if score>=4 else "주의" if score>=2 else "정상"
        results.append({**r,"corp":corp_name,"risk":risk,"risk_score":score,"reasons":reasons,"amount_clean":amt})
    return results

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
